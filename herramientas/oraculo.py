"""Genera los archivos dorados: ejecuta la implementación Python de referencia
sobre el corpus y vuelca lo que devuelve.

Es una herramienta de desarrollo de ESTE repositorio, no del otro. Vive aquí a
propósito: el repositorio de referencia quedó congelado —se consulta, no se
modifica— y ampliar el corpus al portar cada módulo no puede exigir un commit
allá.

    Referencia (solo lectura): github.com/haroldstyven/Formateador

Ese repositorio conserva una copia de este script, de cuando el oráculo todavía
vivía ahí. La copia canónica es esta.

La suite no ejecuta nada de esto: compara contra los JSON ya versionados. Este
script solo se corre al ampliar el corpus.

Uso:

    python herramientas/oraculo.py --modulo valores --generar-corpus
    python herramientas/oraculo.py --modulo valores > herramientas/salida-python-valores.json

Espera el repositorio de referencia en `../Formateador-Banner`; para otra ruta,
la variable de entorno FORMATEADOR_REFERENCIA.

CUANDO LAS DOS IMPLEMENTACIONES DIFIERAN. La referencia está congelada, así que
ya no se le puede aplicar una corrección. La regla pasa a ser:

  1. Decidir cuál tiene razón. El oráculo es la referencia, no la autoridad.
  2. Si tiene razón el port, se regenera el dorado y **el commit explica por
     qué el valor nuevo es el correcto**. Un dorado que cambia sin esa
     explicación es un test desactivado en silencio.
  3. Si tiene razón la referencia, se corrige el port.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from decimal import Decimal
from pathlib import Path

AQUI = Path(__file__).resolve().parent


def _localizar_referencia() -> Path:
    raiz = Path(
        os.environ.get("FORMATEADOR_REFERENCIA")
        or AQUI.parent.parent / "Formateador-Banner"
    ).resolve()
    if not (raiz / "formateador" / "redondeo.py").exists():
        raise SystemExit(
            f"No encuentro la implementación de referencia en {raiz}.\n"
            "Clona github.com/haroldstyven/Formateador junto a este repositorio, "
            "o indica la ruta con --referencia / FORMATEADOR_REFERENCIA."
        )
    return raiz


# Se resuelve al importar, porque los `from formateador...` de abajo dependen
# de ella. Por eso la ruta se configura por entorno y no por argumento.
RAIZ = _localizar_referencia()
sys.path.insert(0, str(RAIZ))

from formateador.redondeo import ValorNoDecimal, formatear  # noqa: E402
from formateador.valores import ConfigValores, interpretar  # noqa: E402
from formateador.mapeo import (  # noqa: E402
    CatalogoAlias,
    mapear,
    puntaje_encabezado,
)
from openpyxl import load_workbook  # noqa: E402

from formateador.lectura import Tabla  # noqa: E402
from formateador.flujo import Analisis, FilaAnalizada  # noqa: E402
from formateador.plantilla import (  # noqa: E402
    EsquemaBanner,
    FilaPlantilla,
    Plantilla,
    cruzar,
)

CORPUS = AQUI / "corpus.json"


def generar_corpus() -> list[dict[str, object]]:
    """Construye el corpus compartido. Se ejecuta una vez y se versiona."""
    casos: list[dict[str, object]] = []

    # Los 50 valores de frontera, como texto y como número.
    for i in range(501):
        texto = f"{Decimal(i) / 100:.2f}"
        if texto.endswith("5"):
            casos.append({"tipo": "texto", "valor": texto})
            casos.append({"tipo": "numero", "valor": float(texto)})

    # Toda la escala ya formateada: verifica la propiedad no-op.
    for i in range(51):
        casos.append({"tipo": "texto", "valor": f"{Decimal(i) / 10:.1f}"})

    # Salidas típicas de una fórmula de promedio ponderado.
    for texto in [
        "4.283333333333333",
        "3.9666666666666668",
        "2.9499999999999997",
        "2.9500000000000002",
        "0.049999999999999996",
        "4.9500000000000002",
        "3.0000000000000004",
    ]:
        casos.append({"tipo": "texto", "valor": texto})
        casos.append({"tipo": "numero", "valor": float(texto)})

    # Enteros y formas alternativas del mismo número.
    for texto in ["0", "3", "5", "3.", ".5", "+4.5", "4.50", "4.5000", " 4.5 ", "4.5e0", "45e-1"]:
        casos.append({"tipo": "texto", "valor": texto})

    # Basura que tiene que ser rechazada igual en las dos implementaciones.
    for texto in ["", "  ", "NP", "4,5", "nan", "inf", "-inf", "1_0", "0x1f", "ver acta", "N/A", "--", "3.0.0"]:
        casos.append({"tipo": "texto", "valor": texto})

    # Fuera de rango: se redondea igual, el rango lo valida otra capa.
    for texto in ["-0.05", "-1.5", "5.05", "9.95"]:
        casos.append({"tipo": "texto", "valor": texto})

    # Cero negativo: Excel lo produce solo y `en_rango` lo acepta, asi que sin
    # normalizar llega hasta `Final Grade`. Lo encontro la prueba diferencial.
    for texto in ["-0.0", "-0", "-0.00", "-0.04", "-0.1"]:
        casos.append({"tipo": "texto", "valor": texto})
    casos.append({"tipo": "numero", "valor": -0.0})

    return casos


def evaluar(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    resultados: list[dict[str, object]] = []
    for caso in casos:
        try:
            resultados.append({"salida": formatear(caso["valor"]), "error": None})
        except ValorNoDecimal:
            resultados.append({"salida": None, "error": "ValorNoDecimal"})
    return resultados


# --------------------------------------------------------------------------
# Módulo `valores`: interpretación de una celda.
#
# Aquí no se compara la prosa de `detalle`, y es deliberado. La representación
# de un Decimal difiere entre Python y decimal.js en los bordes —Decimal("0.0")
# imprime "0.0", decimal.js imprime "0"— y esas cadenas van dentro del mensaje.
# Perseguir igualdad byte a byte en el texto obligaría a deformar una de las
# dos implementaciones para que se parezca a la otra.
#
# Lo que sí se compara es todo lo que decide algo: el estado, el valor emitido,
# el token, y las tres banderas que gobiernan si el archivo se puede generar y
# si la nota entra al diff. El principio §4.4 ya dice que la prosa nunca es la
# fuente de verdad; esto es consistente con eso.
# --------------------------------------------------------------------------

CORPUS_VALORES = AQUI / "corpus-valores.json"

# La config se lee de ESTE repositorio, no del de referencia: es la misma que
# carga `tests/oraculo.test.ts`, y si fueran dos archivos distintos el dorado
# podría producirse con una política y compararse contra otra.
CONFIG_REPO = AQUI.parent / "config" / "valores_no_numericos.json"


def _configs() -> dict[str, ConfigValores]:
    """Las configuraciones con nombre que el corpus puede referenciar."""
    return {
        "vacia": ConfigValores(),
        "repo": ConfigValores.desde_json(CONFIG_REPO),
        "np_reemplazo": ConfigValores.desde_dict(
            {"tokens": {"np": {"accion": "reemplazar", "valor": "0.0"}}}
        ),
        "np_dejar_vacio": ConfigValores.desde_dict(
            {"tokens": {"np": {"accion": "dejar_vacio"}}}
        ),
        "np_descartar": ConfigValores.desde_dict(
            {"tokens": {"np": {"accion": "descartar_fila"}}}
        ),
    }


def generar_corpus_valores() -> list[dict[str, object]]:
    casos: list[dict[str, object]] = []

    def añadir(valor, config="vacia", tipo="texto", formula=False):
        casos.append(
            {"config": config, "tipo": tipo, "valor": valor, "formula": formula}
        )

    # Vacías y fórmulas sin calcular: los dos casos que no se pueden colapsar.
    for v in [None, "", "   ", "\t", "\n  \t"]:
        añadir(v)
    añadir(None, formula=True)
    añadir("4.5", formula=True)

    # Numéricas: la escala completa, la frontera y las salidas de fórmula.
    for i in range(51):
        añadir(f"{Decimal(i) / 10:.1f}")
    for i in range(501):
        texto = f"{Decimal(i) / 100:.2f}"
        if texto.endswith("5"):
            añadir(texto)
            añadir(float(texto), tipo="numero")
    for texto in ["4.283333333333333", "3.9666666666666668", "2.9499999999999997"]:
        añadir(texto)
        añadir(float(texto), tipo="numero")

    # Coma decimal, ambigüedades y formas raras de escribir un número.
    for texto in ["4,5", "4,25", "0,05", "2,95", "1,234.5", "1.234,5", "4,5,6", "4,", ",5"]:
        añadir(texto)
    for texto in ["4.50", "3.", ".5", "+4.5", " 4.5 ", "4.5e0", "45e-1", "0", "5"]:
        añadir(texto)

    # Fuera de rango: no se recorta, se reporta.
    for texto in ["85", "5.1", "-1", "100", "-0.05", "5.05", "9.95"]:
        añadir(texto)

    # Cero negativo: pasa `en_rango` porque -0.0 == 0.0, asi que si no se
    # normaliza sale escrito "-0.0" en la columna de nota.
    for texto in ["-0.0", "-0", "-0.00"]:
        añadir(texto)
    añadir(-0.0, tipo="numero")

    # Tokens administrativos, contra cada configuración con nombre.
    tokens = [
        "NP", "np", "N.P.", " Np ", "n.p.", "N/P",
        "null", "NULL", "None", "NA", "N/A", "-", "--",
        "retiro", "Retirado", "cancelo", "Cancelación", "I", "INC",
        "pendiente", "homologación", "excusa", "No Presentó", "sin nota",
        "ver acta", "texto arbitrario", "?", "0,0,0",
    ]
    for config in ("vacia", "repo", "np_reemplazo", "np_dejar_vacio", "np_descartar"):
        for token in tokens:
            añadir(token, config=config)

    # Booleanos y tipos que no son notas.
    añadir(True, tipo="crudo")
    añadir(False, tipo="crudo")

    return casos


def evaluar_valores(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    configs = _configs()
    resultados: list[dict[str, object]] = []

    for caso in casos:
        config = configs[str(caso["config"])]
        nota = interpretar(
            caso["valor"],
            config,
            formula_sin_calcular=bool(caso.get("formula")),
        )
        resultados.append(
            {
                "estado": nota.estado.value,
                # Un decimal exacto: es lo que se escribe en Banner.
                "valor": None if nota.valor is None else f"{nota.valor:.1f}",
                "token": nota.token,
                "formato_corregido": nota.formato_corregido,
                "requiere_decision": nota.requiere_decision,
                "fue_modificada": nota.fue_modificada,
                "tiene_valor_previo": nota.valor_previo is not None,
            }
        )

    return resultados


# --------------------------------------------------------------------------
# Módulo `mapeo`: qué columna es cuál.
#
# Es el corpus que más importa de los tres, porque `mapeo.ts` no traduce solo
# código nuestro: reimplementa `difflib.SequenceMatcher`. Los puntajes se
# comparan como doubles exactos, sin tolerancia — si el algoritmo quedó "casi
# igual", la diferencia solo se nota en la frontera de 0.90 y 0.78, que es
# justo donde se decide si una columna se elige sola o se pregunta.
#
# Tampoco aquí se compara la prosa de `motivo`: incluye un `f"{x:.2f}"`, y el
# formateo de flotantes en el punto medio no coincide entre Python y JavaScript.
# --------------------------------------------------------------------------

CORPUS_MAPEO = AQUI / "corpus-mapeo.json"

CATALOGO_REPO = AQUI.parent / "config" / "alias_columnas.json"

# Los 13 encabezados de la plantilla oficial de Banner (`plan.md` §2.2).
PLANTILLA_BANNER = [
    "Term Code", "CRN", "Course", "Student ID", "Full Name", "Final Grade",
    "Rolled", "Confidential", "Last Attended Date", "Hours Attended",
    "Incomplete Final Grade", "Extension Date", "Extension Date Constraints",
]

# El archivo real de Humberto (`plan.md` §0.8).
ARCHIVO_REAL = [
    "nombre", "codigo de estudiante", "trabajos", "quizes", "examen",
    "nota definitiva",
]


def generar_corpus_mapeo() -> list[dict[str, object]]:
    filas: list[list[object]] = [
        ARCHIVO_REAL,
        PLANTILLA_BANNER,
        [],
        ["columna a", "columna b"],
        ["codigo", "nota examen", "nota parcial"],
        ["codigo de estudiante", "examen"],
        ["codigo", "nota definitiva", "definitiva"],
        ["codigo de estudiante", "promedio"],
        # Variantes de escritura: acentos, mayúsculas, guiones bajos, erratas.
        ["Código de Estudiante", "NOTA DEFINITIVA"],
        ["CODIGO_DE_ESTUDIANTE", "Nota  Definitiva"],
        ["codigo", "nota_definitiva"],
        ["codigo", "Nota Definitva"],
        ["codigo", "notta definitiva"],
        ["codigo", "nta definitva"],
        ["codigo", "definitivas"],
        ["codigo", "calificacion definitiva"],
        ["codigo", "calificasion final"],
        # Alias débiles: plausibles, nunca seguros.
        ["id", "nota"],
        ["no", "total"],
        ["numero", "final"],
        ["documento", "promedio"],
        # Ambigüedades reales entre dos candidatas fuertes.
        ["cedula", "documento", "nota final", "nota definitiva"],
        ["student id", "final grade", "definitiva"],
        ["codigo", "nombre", "nombres", "nombre completo", "definitiva"],
        # Solo señuelos: no puede resolver la nota.
        ["trabajos", "quizes", "examen", "parcial", "taller"],
        ["primer corte", "segundo corte", "tercer corte"],
        # Ruido, vacíos y tipos que no son texto.
        ["", "   ", "codigo", "definitiva"],
        [None, "codigo de estudiante", "nota definitiva"],
        [1, 2.5, "codigo", "definitiva"],
        ["   nota   definitiva   ", "   codigo   "],
        ["!!!", "???", "codigo", "definitiva"],
        # Encabezados largos, para tocar el camino de la coincidencia difusa.
        ["codigo del estudiante matriculado", "nota definitiva del curso"],
        ["identificacion", "calificacion final del semestre"],
    ]

    # Cada fila del archivo real, aislada: fuerza el caso de una sola columna.
    for encabezado in ARCHIVO_REAL + PLANTILLA_BANNER:
        filas.append([encabezado])

    return [{"encabezados": fila} for fila in filas]


def evaluar_mapeo(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    catalogo = CatalogoAlias.desde_json(CATALOGO_REPO)
    resultados: list[dict[str, object]] = []

    for caso in casos:
        encabezados = list(caso["encabezados"])  # type: ignore[arg-type]
        mapa = mapear(encabezados, catalogo)

        campos: dict[str, object] = {}
        for campo, a in mapa.items():
            campos[campo] = {
                "encabezado": a.encabezado,
                "indice": a.indice,
                "confianza": a.confianza.value,
                "resuelto": a.resuelto,
                "requiere_confirmacion": a.requiere_confirmacion,
                # Puntajes exactos: es lo que valida el port de difflib.
                "candidatas": [
                    {"encabezado": c.encabezado, "indice": c.indice, "puntaje": c.puntaje}
                    for c in a.candidatas
                ],
            }

        resultados.append(
            {
                "campos": campos,
                # Lo usa el lector para encontrar la fila de encabezados.
                "puntaje_encabezado": puntaje_encabezado(encabezados, catalogo),
            }
        )

    return resultados


# --------------------------------------------------------------------------
# Módulo `plantilla`: el cruce por identificador.
#
# Se evalúa `cruzar()` sobre analisis y plantillas armados a mano, sin pasar
# por el lector ni por `analizar_archivo`: lo que se contrasta es la logica del
# cruce, no la lectura del .xlsx. La lectura y la escritura las cubre la suite
# del adaptador, contra el ejemplar anonimizado.
#
# Aqui SI se comparan los motivos de bloqueo, a diferencia de los otros
# modulos: son plantillas de texto con conteos enteros, sin ningun flotante
# formateado, asi que la igualdad literal es alcanzable y vale la pena.
# --------------------------------------------------------------------------

CORPUS_PLANTILLA = AQUI / "corpus-plantilla.json"

IDS = [f"ID-{n:03d}" for n in range(1, 13)]


def _fila(numero: int, codigo: str, valor: object) -> FilaAnalizada:
    return FilaAnalizada(
        numero=numero,
        codigo=codigo,
        nombre=f"Estudiante {numero:03d}",
        nota=interpretar(valor),
    )


def _analisis(filas: list[FilaAnalizada], origen: str | None = "notas.xlsx") -> Analisis:
    tabla = Tabla(
        encabezados=[],
        filas=[],
        origen=Path(origen) if origen else None,
    )
    return Analisis(tabla=tabla, mapa={}, indice_nota=0, filas=filas)


def _plantilla(
    ids: list[str],
    rolled: list[str] | None = None,
    nota_existente: dict[str, str] | None = None,
    origen: str = "Template_Anonimo.xlsx",
) -> Plantilla:
    rolled = rolled or []
    nota_existente = nota_existente or {}
    filas = tuple(
        FilaPlantilla(
            fila=i + 2,
            identificador=ident,
            nombre=f"Anonimo {i + 1}",
            rolled=ident in rolled,
            confidencial=False,
            nota_existente=nota_existente.get(ident, ""),
        )
        for i, ident in enumerate(ids)
    )
    return Plantilla(
        ruta=Path(origen),
        esquema=EsquemaBanner.desde_json(AQUI.parent / "config" / "schema_banner.json"),
        columnas={"Student ID": 4, "Final Grade": 8},
        filas=filas,
        control={"Course": "ANON-101", "Term Code": "202610", "CRN": "12345"},
    )


def generar_corpus_plantilla() -> list[dict[str, object]]:
    """Cada caso declara las notas del docente y la forma de la plantilla."""
    casos: list[dict[str, object]] = []

    def añadir(notas, ids=None, rolled=None, nota_existente=None,
               origen_analisis="notas.xlsx", origen_plantilla="Template_Anonimo.xlsx"):
        casos.append({
            "notas": notas,                       # lista de [codigo, valor]
            "ids": ids if ids is not None else IDS,
            "rolled": rolled or [],
            "nota_existente": nota_existente or {},
            "origen_analisis": origen_analisis,
            "origen_plantilla": origen_plantilla,
        })

    completo = [[i, "4.25"] for i in IDS]

    añadir(completo)
    añadir([[i, "3.5"] for i in IDS])
    añadir([[i, "2.95"] for i in IDS])          # la frontera de aprobacion
    añadir([[i, "4,5"] for i in IDS])           # coma decimal
    añadir([[i, "3.5"] for i in reversed(IDS)])  # orden invertido
    añadir([[i, "3.5"] for i in IDS[:-2]])      # dos estudiantes sin nota
    añadir([])                                   # archivo vacio
    añadir([["ID-999", "5.0"]])                  # solo un intruso
    añadir([[i, "3.5"] for i in IDS] + [["ID-999", "5.0"]])
    añadir([[i, "NP" if k == 3 else "3.5"] for k, i in enumerate(IDS)])
    añadir([[i, "" if k == 7 else "3.5"] for k, i in enumerate(IDS)])
    añadir([[i, "85" if k == 0 else "3.5"] for k, i in enumerate(IDS)])
    añadir([[i.lower().replace("-", " "), "3.5"] for i in IDS])
    añadir([[i, "3.5"] for i in IDS] + [[IDS[0], "1.0"]])   # duplicado
    añadir(completo, rolled=[IDS[2]])
    añadir(completo, rolled=IDS)
    añadir(completo, nota_existente={IDS[0]: "4.0", IDS[1]: "2.0"})
    # Mismo archivo: el docente subio la plantilla ya diligenciada.
    añadir(completo, nota_existente={IDS[0]: "4.0"},
           origen_analisis="Template_Anonimo.xlsx")
    añadir(completo, ids=IDS[:3])                # plantilla mas corta
    añadir([[i, "3.5"] for i in IDS[:3]], ids=IDS[:3])
    añadir([["", "3.5"], [IDS[0], "3.5"]], ids=[IDS[0]])   # codigo vacio

    return casos


def evaluar_plantilla(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    resultados: list[dict[str, object]] = []

    for caso in casos:
        filas = [
            _fila(k + 1, str(codigo), valor)
            for k, (codigo, valor) in enumerate(caso["notas"])  # type: ignore[arg-type]
        ]
        analisis = _analisis(filas, str(caso["origen_analisis"]))
        plantilla = _plantilla(
            list(caso["ids"]),                    # type: ignore[arg-type]
            list(caso["rolled"]),                 # type: ignore[arg-type]
            dict(caso["nota_existente"]),         # type: ignore[arg-type]
            str(caso["origen_plantilla"]),
        )
        cruce = cruzar(analisis, plantilla)

        resultados.append(
            {
                "emparejados": [
                    {
                        "identificador": e.plantilla.identificador,
                        "nota": e.nota_texto,
                        "normalizado": e.identificador_normalizado,
                    }
                    for e in cruce.emparejados
                ],
                "sin_nota": [f.identificador for f in cruce.sin_nota],
                "ya_consolidados": [e.plantilla.identificador for e in cruce.ya_consolidados],
                "sobrantes": [f.codigo for f in cruce.sobrantes],
                "pendientes": [f.codigo for f in cruce.pendientes],
                "sobrescriben_nota": [
                    e.plantilla.identificador for e in cruce.sobrescriben_nota
                ],
                "mismo_archivo": cruce.mismo_archivo,
                "puede_generar": cruce.puede_generar,
                "motivos": cruce.motivos_de_bloqueo,
            }
        )

    return resultados


# --------------------------------------------------------------------------
# Modulo `plantilla-io`: leer y escribir el .xlsx.
#
# AQUI EL DIFERENCIAL NO PUEDE COMPARAR BYTES, y es a proposito. El adaptador
# TypeScript no reabre el libro: parchea el XML dentro del zip, que es la
# solucion que plan.md 2.2 dejo escrita para el caso BL-07b. El objetivo
# declarado es producir bytes DISTINTOS de los de openpyxl, y mas fieles al
# original.
#
# Lo que si tiene que coincidir es el contenido: la plantilla leida, y las
# celdas que se leen de vuelta del archivo generado. Si las dos
# implementaciones producen la misma tabla de datos, la diferencia de
# contenedor es exactamente la mejora que se buscaba.
# --------------------------------------------------------------------------

CORPUS_PLANTILLA_IO = AQUI / "corpus-plantilla-io.json"

FIXTURE = AQUI.parent / "tests" / "fixtures" / "Template_Anonimo.xlsx"


def generar_corpus_plantilla_io() -> list[dict[str, object]]:
    """Cada caso es una nota a escribir en todo el curso, o un cargue parcial."""
    return [
        {"nota": None, "cuantos": 0},        # solo lectura, sin escribir
        {"nota": "4.25", "cuantos": None},   # None = todos
        {"nota": "2.95", "cuantos": None},
        {"nota": "3", "cuantos": None},
        {"nota": "0.05", "cuantos": None},
        {"nota": "5", "cuantos": None},
        {"nota": "4,5", "cuantos": None},
        {"nota": "3.5", "cuantos": 5},       # parcial, con forzar
        {"nota": "3.5", "cuantos": 1},
    ]


def evaluar_plantilla_io(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    import tempfile

    from formateador.plantilla import escribir_en_plantilla, leer_plantilla

    resultados: list[dict[str, object]] = []
    plantilla = leer_plantilla(FIXTURE)

    lectura = {
        "hoja": plantilla.esquema.hoja,
        "columnas": list(plantilla.columnas),
        "curso": plantilla.curso,
        "periodo": plantilla.periodo,
        "crn": plantilla.crn,
        "filas": [
            {
                "fila": f.fila,
                "identificador": f.identificador,
                "nombre": f.nombre,
                "rolled": f.rolled,
                "confidencial": f.confidencial,
                "nota_existente": f.nota_existente,
            }
            for f in plantilla.filas
        ],
    }

    for caso in casos:
        if caso["nota"] is None:
            resultados.append({"lectura": lectura, "celdas": None})
            continue

        cuantos = caso["cuantos"]
        filas_plantilla = (
            plantilla.filas if cuantos is None else plantilla.filas[: int(cuantos)]  # type: ignore[arg-type]
        )
        filas = [
            _fila(i + 1, f.identificador, caso["nota"])
            for i, f in enumerate(filas_plantilla)
        ]
        cruce = cruzar(_analisis(filas), plantilla)

        with tempfile.TemporaryDirectory() as tmp:
            destino = Path(tmp) / "salida.xlsx"
            escribir_en_plantilla(cruce, destino, forzar=True)
            libro = load_workbook(destino, data_only=True)
            hoja = libro[plantilla.esquema.hoja]
            celdas = [
                [
                    "" if hoja.cell(row=r, column=c).value is None
                    else str(hoja.cell(row=r, column=c).value)
                    for c in range(1, hoja.max_column + 1)
                ]
                for r in range(1, hoja.max_row + 1)
            ]

        resultados.append({"lectura": lectura, "celdas": celdas})

    return resultados


# --------------------------------------------------------------------------
# Modulo `lectura`: el lector tolerante.
#
# Las fixtures se generan aqui con openpyxl y se versionan, para que las dos
# implementaciones lean exactamente los mismos bytes. Fabricarlas en cada lado
# no serviria: la mitad de lo que se compara -formatos, formulas sin calcular,
# celdas combinadas- depende de como quedo escrito el archivo.
#
#   python herramientas/oraculo.py --modulo lectura --generar-fixtures
# --------------------------------------------------------------------------

CORPUS_LECTURA = AQUI / "corpus-lectura.json"

FIXTURES = AQUI.parent / "tests" / "fixtures" / "lectura"

ENCABEZADOS = [
    "nombre", "codigo de estudiante", "trabajos", "quizes", "examen",
    "nota definitiva",
]
FILAS_BASE = [
    ["Ana Munoz", "0012345", 4.0, 4.5, 4.2, 4.25],
    ["Luis Pena", "0012346", 3.0, 2.8, 3.1, 2.95],
]


def generar_fixtures() -> list[str]:
    from openpyxl import Workbook

    FIXTURES.mkdir(parents=True, exist_ok=True)
    escritos: list[str] = []

    def guardar(libro, nombre):
        libro.save(FIXTURES / nombre)
        escritos.append(nombre)

    def basico(filas_previas=0, hojas=1, nombre="basico.xlsx"):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Notas"
        for _ in range(filas_previas):
            hoja.append([])
        hoja.append(ENCABEZADOS)
        for fila in FILAS_BASE:
            hoja.append(fila)
        for extra in range(1, hojas):
            libro.create_sheet(f"Hoja{extra}")
        guardar(libro, nombre)

    basico()
    basico(filas_previas=3, nombre="encabezado-fila-4.xlsx")
    basico(hojas=3, nombre="tres-hojas.xlsx")

    # Filas vacias intercaladas.
    libro = Workbook()
    hoja = libro.active
    hoja.append(ENCABEZADOS)
    hoja.append(FILAS_BASE[0])
    hoja.append([])
    hoja.append(FILAS_BASE[1])
    guardar(libro, "huecos.xlsx")

    # Formula sin calcular junto a una celda genuinamente vacia (3.1).
    libro = Workbook()
    hoja = libro.active
    hoja.append(ENCABEZADOS)
    hoja.append(["Ana Munoz", "0012345", 4.0, 4.5, 4.2, "=AVERAGE(C2:E2)"])
    hoja.append(["Luis Pena", "0012346", 3.0, 2.8, 3.1, None])
    guardar(libro, "formulas.xlsx")

    # El formato que oculta precision (3.2).
    for formato, nombre in [
        ("0.0", "formato-un-decimal.xlsx"),
        ("0.00", "formato-dos-decimales.xlsx"),
        ("General", "formato-general.xlsx"),
        ("@", "formato-texto.xlsx"),
    ]:
        libro = Workbook()
        hoja = libro.active
        hoja.append(ENCABEZADOS)
        hoja.append(["Ana", "1", 4.0, 4.5, 4.2, 4.25])
        hoja.cell(row=2, column=6).number_format = formato
        guardar(libro, nombre)

    # Encabezado combinado.
    libro = Workbook()
    hoja = libro.active
    hoja.append(["Consolidado de notas", None, None])
    hoja.merge_cells("A1:C1")
    hoja.append(["codigo de estudiante", "examen", "nota definitiva"])
    hoja.append(["0012345", 4.2, 4.25])
    guardar(libro, "combinadas.xlsx")

    # Fila de resumen al final.
    libro = Workbook()
    hoja = libro.active
    hoja.append(ENCABEZADOS)
    for fila in FILAS_BASE:
        hoja.append(fila)
    hoja.append([None, "PROMEDIO", None, None, None, 3.6])
    guardar(libro, "resumen.xlsx")

    # Sin encabezado reconocible.
    libro = Workbook()
    hoja = libro.active
    for _ in range(5):
        hoja.append([1, 2, 3])
    guardar(libro, "sin-encabezado.xlsx")

    # Valores raros en la columna de nota.
    libro = Workbook()
    hoja = libro.active
    hoja.append(ENCABEZADOS)
    hoja.append(["A", "1", 4.0, 4.0, 4.0, "NP"])
    hoja.append(["B", "2", 4.0, 4.0, 4.0, ""])
    hoja.append(["C", "3", 4.0, 4.0, 4.0, "4,5"])
    hoja.append(["D", "4", 4.0, 4.0, 4.0, 85])
    hoja.append(["E", "5", 4.0, 4.0, 4.0, 0])
    hoja.append(["F", "6", 4.0, 4.0, 4.0, True])
    guardar(libro, "valores-raros.xlsx")

    # Columnas fantasma a la derecha del encabezado.
    libro = Workbook()
    hoja = libro.active
    hoja.append(ENCABEZADOS + [None, None])
    hoja.append(FILAS_BASE[0] + [None, "basura"])
    guardar(libro, "columnas-fantasma.xlsx")

    # Archivos de texto.
    textos = [
        ("punto-y-coma.csv",
         "codigo de estudiante;nota definitiva\n0012345;4,25\n", "utf-8"),
        ("coma.csv",
         "codigo de estudiante,nota definitiva\n0012345,4.25\n", "utf-8"),
        ("latin1.csv",
         "nombre;codigo de estudiante;nota definitiva\nAna Mu\xf1oz Pe\xf1a;0012345;4.5\n",
         "latin-1"),
        ("utf8-bom.csv",
         "codigo de estudiante;nota definitiva\n0012345;4.25\n", "utf-8-sig"),
        ("comillas.csv",
         'codigo de estudiante;nombre;nota definitiva\n0012345;"Munoz; Ana";4.25\n',
         "utf-8"),
        ("tabulador.txt",
         "codigo de estudiante\tnota definitiva\n0012345\t4.25\n", "utf-8"),
    ]
    for nombre, contenido, codificacion in textos:
        (FIXTURES / nombre).write_text(contenido, encoding=codificacion)
        escritos.append(nombre)

    return escritos


def generar_corpus_lectura() -> list[dict[str, object]]:
    return [
        {"archivo": nombre}
        for nombre in sorted(p.name for p in FIXTURES.iterdir() if p.is_file())
    ]


def evaluar_lectura(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    from formateador.lectura import ArchivoNoSoportado, SinEncabezado, leer
    from formateador.valores import interpretar

    catalogo = CatalogoAlias.desde_json(AQUI.parent / "config" / "alias_columnas.json")
    resultados: list[dict[str, object]] = []

    for caso in casos:
        ruta = FIXTURES / str(caso["archivo"])
        try:
            tabla = leer(ruta, catalogo)
        except (SinEncabezado, ArchivoNoSoportado) as exc:
            resultados.append({"error": type(exc).__name__, "tabla": None})
            continue

        filas = []
        for fila in tabla.filas:
            celdas = []
            for c in fila:
                nota = interpretar(c.valor, formula_sin_calcular=c.formula_sin_calcular)
                celdas.append(
                    {
                        "texto": c.texto(),
                        "formula": c.formula,
                        "formato": c.formato,
                        "vacia": c.vacia,
                        "formula_sin_calcular": c.formula_sin_calcular,
                        "oculta_precision": c.oculta_precision,
                        # Lo que de verdad llega a Banner: la prueba mas dura.
                        "estado": nota.estado.value,
                        "nota": None if nota.valor is None else f"{nota.valor:.1f}",
                    }
                )
            filas.append(celdas)

        resultados.append(
            {
                "error": None,
                "tabla": {
                    "encabezados": tabla.encabezados,
                    "hoja": tabla.hoja,
                    "fila_encabezado": tabla.fila_encabezado,
                    "incidencias": tabla.incidencias,
                    "filas": filas,
                },
            }
        )

    return resultados


# --------------------------------------------------------------------------
# Modulo `flujo`: analisis fila por fila y reporte.
#
# Se evalua sobre las fixtures de `tests/fixtures/flujo/`, con varias
# configuraciones de valores no numericos, porque la decision del docente sobre
# un token cambia si el archivo se puede generar o no.
#
# `generado` no se compara: es una marca de tiempo.
# --------------------------------------------------------------------------

CORPUS_FLUJO = AQUI / "corpus-flujo.json"

FIXTURES_FLUJO = AQUI.parent / "tests" / "fixtures" / "flujo"


def generar_fixtures_flujo() -> list[str]:
    from openpyxl import Workbook

    from formateador.plantilla import leer_plantilla

    FIXTURES_FLUJO.mkdir(parents=True, exist_ok=True)
    escritos: list[str] = []

    def archivo(nombre, filas, encabezados=None):
        libro = Workbook()
        hoja = libro.active
        hoja.append(encabezados or ENCABEZADOS)
        for fila in filas:
            hoja.append(fila)
        libro.save(FIXTURES_FLUJO / nombre)
        escritos.append(nombre)

    archivo("limpio.xlsx", [
        ["Ana Munoz", "0012345", 4.0, 4.5, 4.2, 4.283333],
        ["Luis Pena", "0012346", 3.0, 2.8, 3.1, 2.95],
        ["Sara Diaz", "0012347", 3.5, 3.5, 3.5, 3.5],
    ])
    archivo("np.xlsx", [
        ["Ana", "0012345", 4.0, 4.0, 4.0, 4.5],
        ["Luis", "0012346", None, None, None, "NP"],
        ["Sara", "0012347", None, None, None, "NP"],
    ])
    archivo("duplicado.xlsx", [
        ["Ana", "0012345", 4.0, 4.0, 4.0, 4.5],
        ["Otra Ana", "0012345", 4.0, 4.0, 4.0, 3.0],
        ["Luis", "0012346", 4.0, 4.0, 4.0, 3.5],
    ])
    archivo("sin-codigo.xlsx", [
        ["Ana", "0012345", 4.0, 4.0, 4.0, 4.5],
        ["Fantasma", None, 4.0, 4.0, 4.0, 3.0],
    ])
    archivo("fuera-de-rango.xlsx", [["Ana", "0012345", None, None, None, 85]])
    archivo("formula-sin-calcular.xlsx", [
        ["Ana", "0012345", 4.0, 4.5, 4.2, "=AVERAGE(C2:E2)"],
    ])
    # Un curso completo con los identificadores REALES del ejemplar anonimizado:
    # inventarlos aqui probaria el cruce contra datos ficticios.
    plantilla = leer_plantilla(FIXTURE)
    ids = [f.identificador for f in plantilla.filas]

    archivo("curso-completo.xlsx", [
        [f"Estudiante {i:03d}", ident, 4.0, 4.0, 4.0, 4.25]
        for i, ident in enumerate(ids, 1)
    ])
    archivo("curso-incompleto.xlsx", [
        [f"Estudiante {i:03d}", ident, 4.0, 4.0, 4.0, 3.5]
        for i, ident in enumerate(ids[:-2], 1)
    ])
    archivo("curso-con-np.xlsx", [
        [f"Estudiante {i:03d}", ident, 4.0, 4.0, 4.0, "NP" if i == 2 else 3.5]
        for i, ident in enumerate(ids, 1)
    ])

    archivo("todo-junto.xlsx", [
        ["Ana", "0012345", 4.0, 4.5, 4.2, 4.25],
        ["Luis", "0012346", 3.0, 2.8, 3.1, "NP"],
        ["Sara", "0012347", 3.5, 3.5, 3.5, "4,5"],
        ["Juan", "0012345", 3.0, 3.0, 3.0, 3.0],
        ["Nadie", None, 1.0, 1.0, 1.0, 1.0],
        ["Pedro", "0012348", None, None, None, None],
    ])

    return escritos


def generar_corpus_flujo() -> list[dict[str, object]]:
    configs = ["vacia", "repo", "np_reemplazo", "np_dejar_vacio", "np_descartar"]
    casos: list[dict[str, object]] = []
    for nombre in sorted(p.name for p in FIXTURES_FLUJO.iterdir() if p.is_file()):
        for config in configs:
            casos.append({"archivo": nombre, "config": config, "indice_nota": None})
    # La columna forzada por el docente: elegir 'examen' a proposito, que es
    # justo el error que 4.0 existe para impedir cuando NO lo pide una persona.
    casos.append({"archivo": "limpio.xlsx", "config": "vacia", "indice_nota": 4})
    return casos


def evaluar_flujo(casos: list[dict[str, object]]) -> list[dict[str, object]]:
    from formateador.flujo import MapeoIncompleto, analizar
    from formateador.lectura import leer as leer_archivo
    from formateador.reporte import Reporte

    catalogo = CatalogoAlias.desde_json(AQUI.parent / "config" / "alias_columnas.json")
    configs = _configs()
    resultados: list[dict[str, object]] = []

    for caso in casos:
        ruta = FIXTURES_FLUJO / str(caso["archivo"])
        tabla = leer_archivo(ruta, catalogo)
        try:
            analisis = analizar(
                tabla,
                configs[str(caso["config"])],
                catalogo=catalogo,
                indice_nota=caso["indice_nota"],
            )
        except MapeoIncompleto as exc:
            resultados.append({"error": "MapeoIncompleto", "reporte": None})
            continue

        reporte = Reporte(analisis)
        datos = reporte.a_dict()
        datos.pop("generado")

        resultados.append(
            {
                "error": None,
                "reporte": datos,
                "filas": [
                    {
                        "numero": f.numero,
                        "codigo": f.codigo,
                        "nombre": f.nombre,
                        "estado": f.nota.estado.value,
                        "nota": None if f.nota.valor is None else f"{f.nota.valor:.1f}",
                        "bloquea": f.bloquea,
                        "motivo_bloqueo": f.motivo_bloqueo,
                        "problemas": list(f.problemas),
                        "avisos": list(f.avisos),
                    }
                    for f in analisis.filas
                ],
                "pendientes_por_token": {
                    token: [f.numero for f in filas]
                    for token, filas in analisis.pendientes_por_token().items()
                },
                "texto": reporte.a_texto(),
            }
        )

    return resultados


MODULOS = {
    "redondeo": (CORPUS, generar_corpus, evaluar),
    "valores": (CORPUS_VALORES, generar_corpus_valores, evaluar_valores),
    "mapeo": (CORPUS_MAPEO, generar_corpus_mapeo, evaluar_mapeo),
    "plantilla": (CORPUS_PLANTILLA, generar_corpus_plantilla, evaluar_plantilla),
    "plantilla-io": (
        CORPUS_PLANTILLA_IO,
        generar_corpus_plantilla_io,
        evaluar_plantilla_io,
    ),
    "lectura": (CORPUS_LECTURA, generar_corpus_lectura, evaluar_lectura),
    "flujo": (CORPUS_FLUJO, generar_corpus_flujo, evaluar_flujo),
}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--generar-corpus", action="store_true")
    parser.add_argument(
        "--generar-fixtures",
        action="store_true",
        help="solo para --modulo lectura: fabrica los .xlsx y .csv de prueba",
    )
    parser.add_argument("--modulo", choices=sorted(MODULOS), default="redondeo")
    args = parser.parse_args()

    # En Windows, `> archivo.json` hereda la codificación de la consola (cp1252)
    # y parte cualquier encabezado con tilde. Los dorados son UTF-8 siempre.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", newline="\n")

    corpus, generar, evaluador = MODULOS[args.modulo]

    if args.generar_fixtures:
        generadores = {"lectura": generar_fixtures, "flujo": generar_fixtures_flujo}
        generador = generadores.get(args.modulo)
        if generador is None:
            print(
                "--generar-fixtures solo aplica a los modulos "
                f"{sorted(generadores)}",
                file=sys.stderr,
            )
            return 1
        for nombre in generador():
            print(f"  + {nombre}", file=sys.stderr)
        return 0

    if args.generar_corpus:
        corpus.write_text(
            json.dumps(generar(), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"corpus escrito en {corpus}", file=sys.stderr)
        return 0

    if not corpus.exists():
        print(
            f"falta {corpus}. Ejecuta: "
            f"python herramientas/oraculo.py --modulo {args.modulo} --generar-corpus",
            file=sys.stderr,
        )
        return 1

    casos = json.loads(corpus.read_text(encoding="utf-8"))
    json.dump(evaluador(casos), sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
